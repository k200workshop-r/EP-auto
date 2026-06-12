from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COURSE_HEADERS = [
    "#",
    "課程代碼",
    "學年度",
    "醫院別",
    "開課單位",
    "課程名稱",
    "課程類型",
    "排除重複課程",
    "上課時間",
    "開始日期",
    "結束日期",
    "上課時間",
    "上課分鐘數",
    "表單發送時間",
    "課程分類",
    "課程備註",
    "上課地點",
    "該課程授課老師為住院醫師",
    "授課老師",
    "授課老師員編",
    "協同老師",
    "協同老師員編",
    "該課程協同&授課老師皆為同一位住院醫師",
    "學生",
    "輔助教材",
    "職類",
    "計畫類別",
    "訓練計畫",
    "訓練科室",
    "符合/不符合",
]

SUMMARY_HEADERS = COURSE_HEADERS[:-1] + ["A.住院醫師主授課程/B.住院醫師擔任協同老師"]


@dataclass(frozen=True)
class ColumnSettings:
    main_teacher_name: str = "S"
    co_teacher_name: str = "U"


def normalize_name(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    return "" if text.lower() == "nan" else text


def normalize_value(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def column_index(letter: str) -> int:
    cleaned = normalize_value(letter).upper()
    if not re.fullmatch(r"[A-Z]+", cleaned):
        raise ValueError(f"欄位代號不正確：{letter}")
    value = 0
    for char in cleaned:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1


def split_names(value: object) -> list[str]:
    text = normalize_value(value)
    if not text:
        return []
    parts = re.split(r"[、,，;；/／\n\r]+", text)
    return [normalize_name(part) for part in parts if normalize_name(part)]


def first_matching_column(columns: Iterable[object], keywords: tuple[str, ...]) -> object | None:
    normalized = [(column, normalize_name(column)) for column in columns]

    for keyword in keywords:
        key = normalize_name(keyword)
        for column, column_name in normalized:
            if column_name == key:
                return column

    for keyword in keywords:
        key = normalize_name(keyword)
        for column, column_name in normalized:
            if key and key in column_name:
                return column

    return None


def find_column_indices_by_keywords(df: pd.DataFrame, keywords: tuple[str, ...]) -> list[int]:
    normalized_columns = [normalize_name(column) for column in df.columns]
    found: list[int] = []

    for keyword in keywords:
        key = normalize_name(keyword)
        for index, column_name in enumerate(normalized_columns):
            if column_name == key and index not in found:
                found.append(index)

    if found:
        return found

    for keyword in keywords:
        key = normalize_name(keyword)
        for index, column_name in enumerate(normalized_columns):
            if key and key in column_name and index not in found:
                found.append(index)

    return found


def series_by_keywords(df: pd.DataFrame, keywords: tuple[str, ...], occurrence: int = 0) -> pd.Series:
    indices = find_column_indices_by_keywords(df, keywords)
    if indices:
        idx = indices[min(occurrence, len(indices) - 1)]
        return df.iloc[:, idx].astype(object)
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def series_by_letter(df: pd.DataFrame, letter: str) -> pd.Series:
    idx = column_index(letter)
    if idx < len(df.columns):
        return df.iloc[:, idx].astype(object)
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def read_excel_first_sheet(uploaded_file) -> pd.DataFrame:
    return pd.read_excel(uploaded_file, sheet_name=0, dtype=object, engine="openpyxl").astype(object)


def build_teacher_lookup(mapping_df: pd.DataFrame) -> dict[str, str]:
    name_col = first_matching_column(
        mapping_df.columns,
        ("教師姓名", "老師姓名", "姓名", "授課老師", "協同老師", "住院醫師姓名"),
    )
    employee_col = first_matching_column(
        mapping_df.columns,
        ("員編", "員工編號", "教師員編", "醫師員編", "住院醫師員編", "職編", "ID"),
    )

    if name_col is None or employee_col is None:
        raise ValueError("比對用住院醫師名單需包含姓名欄位與員編欄位，例如「姓名」與「員編」。")

    lookup: dict[str, str] = {}
    for _, row in mapping_df.iterrows():
        name = normalize_name(row[name_col])
        employee_id = normalize_value(row[employee_col])
        if name and employee_id:
            lookup[name] = employee_id
    return lookup


def teacher_ids_for_value(value: object, lookup: dict[str, str]) -> str:
    ids = [lookup[name] for name in split_names(value) if name in lookup]
    # 去除同一格內重複員編，但保留原順序
    unique_ids: list[str] = []
    for employee_id in ids:
        if employee_id not in unique_ids:
            unique_ids.append(employee_id)
    return "、".join(unique_ids)


def same_resident_teacher(main_value: object, co_value: object, lookup: dict[str, str]) -> bool:
    main_ids = set(teacher_ids_for_value(main_value, lookup).split("、")) - {""}
    co_ids = set(teacher_ids_for_value(co_value, lookup).split("、")) - {""}
    return bool(main_ids and co_ids and main_ids & co_ids)


def value_contains_teacher(student_value: object, teacher_value: object) -> bool:
    """判斷協同老師是否同時出現在學生欄位。"""
    student_names = set(split_names(student_value))
    teacher_names = set(split_names(teacher_value))
    return bool(student_names and teacher_names and (student_names & teacher_names))


def build_dedupe_key(df: pd.DataFrame) -> pd.Series:
    # 對應原 Excel：I欄 & Q欄 & S欄。僅標示，不刪除資料。
    i_series = series_by_letter(df, "I").map(normalize_value)
    q_series = series_by_letter(df, "Q").map(normalize_value)
    s_series = series_by_letter(df, "S").map(normalize_value)
    return i_series + "|" + q_series + "|" + s_series


def standardize_report(df: pd.DataFrame, lookup: dict[str, str], settings: ColumnSettings) -> pd.DataFrame:
    df = df.copy().astype(object)

    # 先用欄名找，找不到才用預設欄位代號
    main_series = series_by_keywords(df, ("授課老師", "授課教師", "主授課老師"))
    if main_series.map(normalize_value).eq("").all():
        main_series = series_by_letter(df, settings.main_teacher_name)

    co_series = series_by_keywords(df, ("協同老師", "協同教師", "協同授課老師"))
    if co_series.map(normalize_value).eq("").all():
        co_series = series_by_letter(df, settings.co_teacher_name)

    student_series = series_by_keywords(df, ("學生", "學員", "受訓學員", "受訓人員"))
    dedupe_series = build_dedupe_key(df)

    output_columns: list[tuple[str, pd.Series]] = []
    course_time_count = 0

    for header in COURSE_HEADERS:
        if header == "排除重複課程":
            s = dedupe_series
        elif header == "該課程授課老師為住院醫師":
            s = main_series.map(lambda v: "授課老師有員編" if teacher_ids_for_value(v, lookup) else "")
        elif header == "授課老師":
            s = main_series
        elif header == "授課老師員編":
            s = main_series.map(lambda v: teacher_ids_for_value(v, lookup))
        elif header == "協同老師":
            s = co_series
        elif header == "協同老師員編":
            s = co_series.map(lambda v: teacher_ids_for_value(v, lookup))
        elif header == "該課程協同&授課老師皆為同一位住院醫師":
            s = pd.Series(
                ["是" if same_resident_teacher(m, c, lookup) else "" for m, c in zip(main_series, co_series)],
                index=df.index,
                dtype=object,
            )
        elif header == "學生":
            s = student_series
        elif header == "符合/不符合":
            s = pd.Series([""] * len(df), index=df.index, dtype=object)
        elif header == "上課時間":
            # 原格式有兩個「上課時間」，依序取第1個、第2個
            s = series_by_keywords(df, ("上課時間",), occurrence=course_time_count)
            course_time_count += 1
        else:
            s = series_by_keywords(df, (header,))

        if not isinstance(s, pd.Series):
            raise RuntimeError(f"輸出欄位「{header}」不是單一欄，請檢查原始欄位是否重複或格式異常。")

        output_columns.append((header, s.astype(object).reset_index(drop=True)))

    output = pd.concat([series for _, series in output_columns], axis=1)
    output.columns = [header for header, _ in output_columns]

    # 「符合/不符合」欄：主要篩選「協同老師是否同時為學生」。
    student_not_blank = output["學生"].map(normalize_value).ne("")
    co_teacher_is_student = pd.Series(
        [
            value_contains_teacher(student, co_teacher)
            for student, co_teacher in zip(output["學生"], output["協同老師"])
        ],
        index=output.index,
        dtype=bool,
    )

    output["符合/不符合"] = "符合"
    output.loc[~student_not_blank, "符合/不符合"] = "不符合(學生空白)"
    output.loc[student_not_blank & co_teacher_is_student, "符合/不符合"] = "不符合(協同老師同時為學生)"

    return output.astype(object)


def build_summary(course_df: pd.DataFrame) -> pd.DataFrame:
    """彙整：先排除「不符合」，再依 A 類優先於 B 類分類。"""
    df = course_df.copy().astype(object)

    eligible = df["符合/不符合"].map(normalize_value).eq("符合")
    main_is_resident = df["授課老師員編"].map(normalize_value).ne("")
    co_is_resident = df["協同老師員編"].map(normalize_value).ne("")

    main_mask = eligible & main_is_resident
    co_mask = eligible & ~main_mask & co_is_resident
    selected_mask = main_mask | co_mask

    summary_base = df.loc[selected_mask, COURSE_HEADERS[:-1]].copy()
    summary_base.columns = SUMMARY_HEADERS[:-1]

    category = pd.Series("", index=df.index, dtype=object)
    category.loc[main_mask] = "A.住院醫師主授課程"
    category.loc[co_mask] = "B.住院醫師擔任協同老師"
    summary_base["A.住院醫師主授課程/B.住院醫師擔任協同老師"] = category.loc[summary_base.index].values

    # 彙整重新編號；課程總表不重新編號，保留原始資料概念。
    summary_base["#"] = range(1, len(summary_base) + 1)

    return summary_base.reset_index(drop=True).astype(object)


def autosize_and_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(text), 40))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 35))


def write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(["" if pd.isna(value) else value for value in row])
    autosize_and_style(ws)


def process_report(report_file, mapping_file) -> tuple[io.BytesIO, int, int, int, int]:
    report_df = read_excel_first_sheet(report_file)
    mapping_df = read_excel_first_sheet(mapping_file)
    lookup = build_teacher_lookup(mapping_df)

    course_df = standardize_report(report_df, lookup, ColumnSettings())
    summary_df = build_summary(course_df)

    wb = Workbook()
    ws_course = wb.active
    ws_course.title = "課程總表"
    write_dataframe(ws_course, course_df)

    ws_summary = wb.create_sheet("彙整")
    write_dataframe(ws_summary, summary_df)

    ws_lookup = wb.create_sheet("比對用")
    lookup_df = pd.DataFrame([{"姓名": name, "員編": employee_id} for name, employee_id in lookup.items()])
    write_dataframe(ws_lookup, lookup_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    blank_student_rows = int((course_df["符合/不符合"] == "不符合(學生空白)").sum())
    co_teacher_student_rows = int((course_df["符合/不符合"] == "不符合(協同老師同時為學生)").sum())

    return output, len(course_df), len(summary_df), blank_student_rows, co_teacher_student_rows


st.set_page_config(page_title="RAST 自動化報表", page_icon="📊", layout="centered")
st.title("RAST 自動化報表")
st.caption("完整保留課程總表原始資料；彙整排除學生空白與協同老師同時為學生者，再計入住院醫師授課/協同。")

report_file = st.file_uploader("上傳原始 RAST Excel 報表", type=["xlsx"])
mapping_file = st.file_uploader("上傳比對用住院醫師名單", type=["xlsx"])

st.info("規則：符合/不符合欄主要用於排除「學生空白」與「協同老師同時為學生」；彙整再依授課老師/協同老師員編判斷 A/B 類。")

if st.button("產生整理後 RAST 報表", type="primary"):
    if report_file is None or mapping_file is None:
        st.error("請同時上傳原始 RAST Excel 報表與比對用住院醫師名單。")
    else:
        try:
            output, course_rows, summary_rows, blank_student_rows, co_teacher_student_rows = process_report(
                report_file,
                mapping_file,
            )
            st.success(
                f"完成：課程總表 {course_rows} 筆；彙整 {summary_rows} 筆；"
                f"學生空白未納入 {blank_student_rows} 筆；"
                f"協同老師同時為學生未納入 {co_teacher_student_rows} 筆。"
            )
            st.download_button(
                label="下載整理後 RAST 報表",
                data=output,
                file_name=f"整理後RAST報表_課程總表{course_rows}筆_彙整{summary_rows}筆.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(f"處理失敗：{exc}")
